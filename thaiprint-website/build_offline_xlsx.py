"""
ไฟล์ Excel คำนวณราคาออฟไลน์ — Thai Print Inter Group
ชีต: Prices (ราคาแก้ได้) / Offset (auto ตัด4-ตัด2 + งานหลังพิมพ์) / Inkjet / Quote (ใบเสนอราคาพร้อมพิมพ์ A4)
สูตร/ราคาตรงกับเว็บ: เผื่อเสีย 300 ใบ/ด้าน, กระดาษ×แกรม×markup 1.20, สติ๊กเกอร์ ppc รวม markup, W&T/Sheetwise
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

TH="TH Sarabun New"
HFILL=PatternFill("solid",fgColor="1F4E79"); SEC=PatternFill("solid",fgColor="D9E1F2")
INP=PatternFill("solid",fgColor="FFF2CC"); RES=PatternFill("solid",fgColor="C6EFCE")
GREY=PatternFill("solid",fgColor="F2F2F2"); ORANGE=PatternFill("solid",fgColor="F7941D")
Hf=Font(name=TH,size=14,bold=True,color="FFFFFF"); B=Font(name=TH,size=13,bold=True)
Nf=Font(name=TH,size=13); Sm=Font(name=TH,size=11,color="808080")
Big=Font(name=TH,size=20,bold=True,color="C0504D"); Logo=Font(name=TH,size=22,bold=True,color="FFFFFF")
CoName=Font(name=TH,size=16,bold=True,color="1F4E79")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
NONE="— ไม่เลือก —"

def sc(ws,cell,val,font=Nf,fill=None,align=None,border=None,numfmt=None):
    c=ws[cell]; c.value=val; c.font=font
    if fill:c.fill=fill
    if align:c.alignment=Alignment(horizontal=align,vertical="center")
    if border:c.border=border
    if numfmt:c.number_format=numfmt
    return c

wb=Workbook()

# ================= PRICES =================
p=wb.active; p.title="Prices"; p.sheet_view.showGridLines=False
sc(p,"A1","ตารางราคา (แก้ไขได้ — ช่องเหลืองปรับได้)",Hf,HFILL); p.merge_cells("A1:E1")
sc(p,"A3","Markup กระดาษ"); sc(p,"B3",1.20,Nf,INP)
sc(p,"A4","เผื่อเสีย/ด้าน (ใบพิมพ์)"); sc(p,"B4",300,Nf,INP)
sc(p,"A5","ค่าเพลท ตัด4 (บ./สี)"); sc(p,"B5",300,Nf,INP)
sc(p,"A6","ค่าเพลท ตัด2 (บ./สี)"); sc(p,"B6",600,Nf,INP)
sc(p,"A7","VAT (%)"); sc(p,"B7",7,Nf,INP)
sc(p,"A9","เรตค่าพิมพ์ (บ./สี)",B,SEC); p.merge_cells("A9:D9")
for i,h in enumerate(["คีย์","เหมา/สี","ช่วงเหมา(ใบ)","เกิน(บ./ใบ)"]): sc(p,f"{'ABCD'[i]}10",h,B,GREY,border=BORD)
for r,(k,fl,up,ov) in enumerate([("cut4_conventional",900,10000,0.10),("cut4_uv",1500,1000,1.00),
                                 ("cut2_conventional",1200,10000,0.20),("cut2_uv",2000,10000,1.50)],start=11):
    sc(p,f"A{r}",k,Nf,border=BORD); sc(p,f"B{r}",fl,Nf,INP,border=BORD)
    sc(p,f"C{r}",up,Nf,INP,border=BORD); sc(p,f"D{r}",ov,Nf,INP,border=BORD,numfmt="0.00")
sc(p,"A17","วัสดุ",B,SEC); p.merge_cells("A17:E17")
for i,h in enumerate(["ชื่อวัสดุ","ชนิด","บ./กก.","บ./ตร.ซม.","หมึก"]): sc(p,f"{'ABCDE'[i]}18",h,B,GREY,border=BORD)
materials=[("อาร์ตมัน","paper",32,None,"conventional"),("อาร์ตด้าน","paper",32,None,"conventional"),
 ("อาร์ตการ์ด 2 หน้า","paper",34,None,"conventional"),("อาร์ตการ์ด 1 หน้า (ไอวอรี่)","paper",32,None,"conventional"),
 ("กระดาษการ์ดขาว","paper",34.5,None,"conventional"),("กระดาษปอนด์","paper",30,None,"conventional"),
 ("กล่องแป้งหลังเทา","paper",23,None,"conventional"),("ไวท์คราฟท์","paper",40,None,"conventional"),
 ("สติ๊กเกอร์กระดาษ","sticker",None,0.002911,"conventional"),("สติ๊กเกอร์ PVC","sticker",None,0.004528,"uv"),
 ("สติ๊กเกอร์ PP ขาวเงา","sticker",None,0.003720,"uv"),("สติ๊กเกอร์ PP ขาวด้าน","sticker",None,0.003922,"uv"),
 ("สติ๊กเกอร์ PP ขาวเงา หลังขาว","sticker",None,0.002385,"uv"),("สติ๊กเกอร์ PP ขาวด้าน หลังขาว","sticker",None,0.002102,"uv"),
 ("สติ๊กเกอร์ PET","sticker",None,0.005175,"uv"),("สติ๊กเกอร์ฟอยล์","sticker",None,0.007278,"uv")]
ms=19
for r,(nm,ty,ppk,ppc,ink) in enumerate(materials,start=ms):
    sc(p,f"A{r}",nm,Nf,border=BORD); sc(p,f"B{r}",ty,Nf,border=BORD)
    sc(p,f"C{r}",ppk if ppk is not None else "",Nf,INP,border=BORD)
    sc(p,f"D{r}",ppc if ppc is not None else "",Nf,INP,border=BORD,numfmt="0.000000"); sc(p,f"E{r}",ink,Nf,border=BORD)
me=ms+len(materials)-1
# ---- inkjet materials table (ตรงกับ INKJET registry ในเว็บ index.html/admin.html) ----
ij_sec = me+3
sc(p,f"A{ij_sec}","อิงค์เจ็ท — วัสดุ & เรตราคา (บ./ตร.ม.)",B,SEC); p.merge_cells(f"A{ij_sec}:K{ij_sec}")
ij_hdr = ij_sec+1
ij_cols=["ชื่อวัสดุ","โหมด","≤ตร.ม.","720(เล็ก)","1440(เล็ก)","2880(เล็ก)","720(ใหญ่)","1440(ใหญ่)","2880(ใหญ่)","flatปกติ","flatด่วน"]
for i,h in enumerate(ij_cols): sc(p,f"{chr(65+i)}{ij_hdr}",h,B,GREY,border=BORD)
# (name, mode, tierMax, s720,s1440,s2880, L720,L1440,L2880, flatNormal, flatRush) — quote=ราคาว่าง
inkjet_rows=[
 ("ไวนิล Frontlit (ทึบแสง)","volume",10,180,220,None,120,150,None,None,None),
 ("ไวนิล Backlit (โปร่งแสง)","quote",None,None,None,None,None,None,None,None,None),
 ("ไวนิล Blockout (กันแสงทะลุ)","quote",None,None,None,None,None,None,None,None,None),
 ("สติ๊กเกอร์ PVC ขาว (Outdoor)","volume",10,150,190,None,110,150,None,None,None),
 ("สติ๊กเกอร์ PVC ใส (Clear)","quote",None,None,None,None,None,None,None,None,None),
 ("สติ๊กเกอร์ Seethrough (One-Way Vision)","quote",None,None,None,None,None,None,None,None,None),
 ("สติ๊กเกอร์ 3M (พรีเมียม/ถอดได้)","quote",None,None,None,None,None,None,None,None,None),
 ("สติ๊กเกอร์ Backlit (ป้ายไฟ)","quote",None,None,None,None,None,None,None,None,None),
 ("PP Sticker (สังเคราะห์)","quote",None,None,None,None,None,None,None,None,None),
 ("โฟโต้เปเปอร์ (Glossy Photo)","flat",None,None,None,None,None,None,None,250,350),
 ("PP Paper (เคลือบ กันน้ำ)","quote",None,None,None,None,None,None,None,None,None),
 ("Matte / Satin / Fine Art Paper","quote",None,None,None,None,None,None,None,None,None),
 ("Blue Back Paper (หลังน้ำเงิน)","quote",None,None,None,None,None,None,None,None,None),
 ("Film Backlit (ฟิล์มป้ายไฟ)","quote",None,None,None,None,None,None,None,None,None),
 ("Clear / Reflective / Floor Film","quote",None,None,None,None,None,None,None,None,None),
 ("ผ้า IT / ผ้าไวนิล (Backdrop)","flat",None,None,None,None,None,None,None,300,420),
 ("ผ้า Tension / Mesh","quote",None,None,None,None,None,None,None,None,None),
 ("Dye-Sub Fabric (ซับลิเมชัน)","quote",None,None,None,None,None,None,None,None,None),
 ("ผ้าแคนวาส (Canvas)","flat",None,None,None,None,None,None,None,400,550),
 ("Wallpaper (วอลเปเปอร์)","quote",None,None,None,None,None,None,None,None,None),
 ("Magnetic (แม่เหล็ก)","quote",None,None,None,None,None,None,None,None,None),
 ("Tarpaulin (ผ้าใบกันน้ำ)","quote",None,None,None,None,None,None,None,None,None),
 ("Rigid Board (พิมพ์ UV ตรง)","quote",None,None,None,None,None,None,None,None,None),
 ("Banner Mesh (ตาข่ายโฆษณา)","quote",None,None,None,None,None,None,None,None,None),
]
ij_start = ij_hdr+1
for r,row in enumerate(inkjet_rows,start=ij_start):
    for i,val in enumerate(row):
        fill = INP if (i>=2 and val is not None) else None
        nf = "0.00" if (i>=3 and isinstance(val,(int,float))) else None
        sc(p,f"{chr(65+i)}{r}", val if val is not None else "", Nf, fill, border=BORD, numfmt=nf)
ij_end = ij_start+len(inkjet_rows)-1
IJ = f"Prices!$A${ij_start}:$K${ij_end}"

# finishing (ใต้ตารางอิงค์เจ็ท)
fin_sec = ij_end+2
sc(p,f"A{fin_sec}","งานหลังพิมพ์ (ยืนยันราคากับโรงงาน)",B,SEC); p.merge_cells(f"A{fin_sec}:D{fin_sec}")
for i,h in enumerate(["ชื่อ","ราคา","หน่วย","ขั้นต่ำ(บ.)"]): sc(p,f"{'ABCD'[i]}{fin_sec+1}",h,B,GREY,border=BORD)
fins=[(NONE,0,"ชิ้น",0),("เคลือบเงา",1.0,"ใบ",300),("เคลือบด้าน",1.0,"ใบ",300),("Spot UV",2.0,"ใบ",500),
 ("พับ",0.10,"ชิ้น",200),("ไดคัท",0.30,"ชิ้น",500),("ปั๊มนูน",0.50,"ชิ้น",500),
 ("ปั๊มเคทอง/เงิน",0.60,"ชิ้น",600),("รันนัมเบอร์",0.15,"ชิ้น",200),("ปรุฉีก",0.10,"ชิ้น",200)]
fs=fin_sec+2
for r,(nm,rate,unit,mn) in enumerate(fins,start=fs):
    sc(p,f"A{r}",nm,Nf,border=BORD); sc(p,f"B{r}",rate,Nf,INP,border=BORD,numfmt="0.00")
    sc(p,f"C{r}",unit,Nf,border=BORD); sc(p,f"D{r}",mn,Nf,INP,border=BORD)
fe=fs+len(fins)-1
for col,w in {"A":34,"B":13,"C":13,"D":13,"E":16,"F":11,"G":11,"H":11,"I":11,"J":11,"K":11}.items(): p.column_dimensions[col].width=w

MAT=f"Prices!$A${ms}:$E${me}"; FIN=f"Prices!$A${fs}:$D${fe}"

# ================= OFFSET =================
o=wb.create_sheet("Offset"); o.sheet_view.showGridLines=False
sc(o,"A1","คำนวณราคางานพิมพ์ Offset",Hf,HFILL); o.merge_cells("A1:F1")
sc(o,"A2","กรอกช่องเหลือง → ระบบเลือกเครื่องถูกกว่า (ตัด4/ตัด2) + งานหลังพิมพ์",Sm); o.merge_cells("A2:F2")
for la,lt,ca,cv in [("A4","วัสดุ","B4","สติ๊กเกอร์ PVC"),("A5","แกรม (เฉพาะกระดาษ)","B5",128),
  ("A6","กว้าง (ซม.)","B6",10),("A7","สูง (ซม.)","B7",10),("A8","สีด้านหน้า","B8",4),
  ("A9","สีด้านหลัง (0=หน้าเดียว)","B9",0),("A10","วิธีพิมพ์ 2 ด้าน","B10","กลับหน้าในตัว"),
  ("A11","จำนวน (ชิ้น)","B11",10000)]:
    sc(o,la,lt,B); sc(o,ca,cv,Nf,INP,border=BORD)
dv=DataValidation(type="list",formula1=f"Prices!$A${ms}:$A${me}"); o.add_data_validation(dv); dv.add(o["B4"])
dv2=DataValidation(type="list",formula1='"พิมพ์หน้าเดียว,กลับหน้าในตัว,Sheetwise"'); o.add_data_validation(dv2); dv2.add(o["B10"])
sc(o,"A14","ชนิดวัสดุ"); sc(o,"B14",f'=VLOOKUP(B4,{MAT},2,FALSE)')
sc(o,"A15","หมึก"); sc(o,"B15",f'=IF(B14="sticker",VLOOKUP(B4,{MAT},5,FALSE),"conventional")')
sc(o,"D17","รายละเอียด",B,SEC); sc(o,"E17","ตัด 4 (MO 65×48)",B,SEC); sc(o,"F17","ตัด 2 (SM74 74×53)",B,SEC)
labels={18:"n-up (ชิ้น/ใบ)",19:"สี (เพลท)",20:"ด้านพิมพ์",21:"ชิ้นสำเร็จ/ใบ",22:"ใบขั้นต่ำ",
 23:"เผื่อเสีย(ใบ)",24:"ใบพิมพ์รวม",25:"ค่าเพลท",26:"ค่าพิมพ์",27:"ค่ากระดาษ/วัสดุ",28:"รวม (ฐาน)"}
for r,t in labels.items(): sc(o,f"D{r}",t,Nf,border=BORD)
PW={"E":65,"F":74}; PH={"E":48,"F":53}
for col in ("E","F"):
    pw,ph=PW[col],PH[col]; plate="$B$5" if col=="E" else "$B$6"; rk="cut4_" if col=="E" else "cut2_"
    sc(o,f"{col}18",f'=MAX(FLOOR(({pw}-1.2)/($B$6+0.6),1)*FLOOR(({ph}-0.7)/($B$7+0.6),1),FLOOR(({pw}-1.2)/($B$7+0.6),1)*FLOOR(({ph}-0.7)/($B$6+0.6),1))',Nf,border=BORD)
    sc(o,f"{col}19",'=IF($B$9=0,$B$8,IF($B$10="กลับหน้าในตัว",MAX($B$8,$B$9),$B$8+$B$9))',Nf,border=BORD)
    sc(o,f"{col}20",'=IF($B$9=0,1,2)',Nf,border=BORD)
    sc(o,f"{col}21",f'=IF(AND($B$9>0,$B$10="กลับหน้าในตัว"),FLOOR({col}18/2,1),{col}18)',Nf,border=BORD)
    sc(o,f"{col}22",f'=CEILING($B$11/{col}21,1)',Nf,border=BORD)
    sc(o,f"{col}23",f'=Prices!$B$4*{col}20',Nf,border=BORD)
    sc(o,f"{col}24",f'={col}22+{col}23',Nf,border=BORD)
    sc(o,f"{col}25",f'={plate}*{col}19',Nf,border=BORD,numfmt="#,##0.00")
    fl=f'INDEX(Prices!$B$11:$B$14,MATCH("{rk}"&$B$15,Prices!$A$11:$A$14,0))'
    up=f'INDEX(Prices!$C$11:$C$14,MATCH("{rk}"&$B$15,Prices!$A$11:$A$14,0))'
    ov=f'INDEX(Prices!$D$11:$D$14,MATCH("{rk}"&$B$15,Prices!$A$11:$A$14,0))'
    sc(o,f"{col}26",f'=({fl}+IF({col}24>{up},({col}24-{up})*{ov},0))*{col}19',Nf,border=BORD,numfmt="#,##0.00")
    ppk=f'VLOOKUP($B$4,{MAT},3,FALSE)'; ppc=f'VLOOKUP($B$4,{MAT},4,FALSE)'
    sc(o,f"{col}27",f'=IF($B$14="paper",({pw}*{ph})*$B$5*{ppk}/10000000*Prices!$B$3*{col}24,{ppc}*($B$6*$B$7)*($B$11+{col}23*{col}21))',Nf,border=BORD,numfmt="#,##0.00")
    sc(o,f"{col}28",f'={col}25+{col}26+{col}27',B,border=BORD,numfmt="#,##0.00")
# best sheets/machine
sc(o,"A30","ใบพิมพ์ (เครื่องที่เลือก)"); sc(o,"B30","=IF(E28<=F28,E24,F24)",Nf)
# finishing selection
sc(o,"A32","งานหลังพิมพ์ (เลือกได้)",B,SEC); o.merge_cells("A32:C32")
sc(o,"A33","รายการ",B,GREY,border=BORD); sc(o,"B33","",B,GREY,border=BORD); sc(o,"C33","ราคา",B,GREY,border=BORD)
dvf=DataValidation(type="list",formula1=f"Prices!$A${fs}:$A${fe}"); o.add_data_validation(dvf)
for r in range(34,38):
    sc(o,f"A{r}",f"งานหลังพิมพ์ {r-33}",Nf,border=BORD)
    sc(o,f"B{r}",NONE,Nf,INP,border=BORD); dvf.add(o[f"B{r}"])
    o.merge_cells(f"A{r}:B{r}")
    fc=(f'=IFERROR(IF($B{r}="{NONE}",0,MAX(VLOOKUP($B{r},{FIN},4,FALSE),'
        f'VLOOKUP($B{r},{FIN},2,FALSE)*IF(VLOOKUP($B{r},{FIN},3,FALSE)="ใบ",$B$30,$B$11))),0)')
    sc(o,f"C{r}",fc,Nf,border=BORD,numfmt="#,##0.00")
sc(o,"A38","อื่น ๆ เพิ่มเติม (บาท)",Nf,border=BORD); o.merge_cells("A38:B38"); sc(o,"C38",0,Nf,INP,border=BORD,numfmt="#,##0.00")
sc(o,"A39","รวมงานหลังพิมพ์",B,border=BORD); o.merge_cells("A39:B39"); sc(o,"C39","=SUM(C34:C38)",B,border=BORD,numfmt="#,##0.00")
# summary
sc(o,"A41","แนะนำเครื่อง",B); sc(o,"B41",'=IF(E28<=F28,"ตัด 4 (MO)","ตัด 2 (SM74)")',B)
sc(o,"A42","ราคาฐาน (เพลท+พิมพ์+วัสดุ)"); sc(o,"B42","=MIN(E28,F28)",Nf,numfmt="#,##0.00")
sc(o,"A43","ค่างานหลังพิมพ์"); sc(o,"B43","=C39",Nf,numfmt="#,##0.00")
sc(o,"A44","ราคารวม (บาท)",B); sc(o,"B44","=B42+B43",Big,RES,numfmt="฿#,##0.00")
sc(o,"A45","ราคาต่อชิ้น"); sc(o,"B45","=(B42+B43)/B11",Nf,numfmt="฿#,##0.0000")
sc(o,"A47","หมายเหตุ: กลับหน้าในตัว (W&T) ใช้เมื่อสีหน้า=สีหลัง; สีต่างกันเลือก Sheetwise · ราคายังไม่รวม VAT",Sm); o.merge_cells("A47:F47")
for col,w in {"A":26,"B":18,"C":14,"D":18,"E":16,"F":16}.items(): o.column_dimensions[col].width=w

# ================= INKJET =================
j=wb.create_sheet("Inkjet"); j.sheet_view.showGridLines=False
sc(j,"A1","คำนวณราคางานอิงค์เจ็ท",Hf,HFILL); j.merge_cells("A1:D1")
for la,lt,ca,cv in [("A3","วัสดุ","B3","ไวนิล Frontlit (ทึบแสง)"),("A4","ความละเอียด (dpi)","B4",1440),
  ("A5","กว้าง (เมตร)","B5",3),("A6","สูง (เมตร)","B6",1),("A7","จำนวน (ป้าย)","B7",5),("A8","งานด่วน","B8","ไม่")]:
    sc(j,la,lt,B); sc(j,ca,cv,Nf,INP,border=BORD)
# วัสดุเลือกจากตารางอิงค์เจ็ทใน Prices (24 รายการ) — ครอบคลุมทุกวัสดุ/ความละเอียด
dvm=DataValidation(type="list",formula1=f"Prices!$A${ij_start}:$A${ij_end}"); j.add_data_validation(dvm); dvm.add(j["B3"])
dvd=DataValidation(type="list",formula1='"720,1440,2880"'); j.add_data_validation(dvd); dvd.add(j["B4"])
dvr=DataValidation(type="list",formula1='"ใช่,ไม่"'); j.add_data_validation(dvr); dvr.add(j["B8"])
sc(j,"A10","พื้นที่รวม (ตร.ม.)"); sc(j,"B10","=B5*B6*B7",Nf,numfmt="0.00")
sc(j,"A11","โหมดราคา"); sc(j,"B11",f'=IFERROR(VLOOKUP(B3,{IJ},2,FALSE),"quote")',Sm)
sc(j,"A12","เรต (บ./ตร.ม.)"); sc(j,"B12",
 f'=IFERROR(IF(B11="quote","สอบถามราคา",'
 f'IF(B11="flat",IF(B8="ใช่",VLOOKUP(B3,{IJ},11,FALSE),VLOOKUP(B3,{IJ},10,FALSE)),'
 f'IF(B10<=VLOOKUP(B3,{IJ},3,FALSE),'
 f'CHOOSE(MATCH(--B4,{{720,1440,2880}},0),VLOOKUP(B3,{IJ},4,FALSE),VLOOKUP(B3,{IJ},5,FALSE),VLOOKUP(B3,{IJ},6,FALSE)),'
 f'CHOOSE(MATCH(--B4,{{720,1440,2880}},0),VLOOKUP(B3,{IJ},7,FALSE),VLOOKUP(B3,{IJ},8,FALSE),VLOOKUP(B3,{IJ},9,FALSE))))),"สอบถามราคา")',
 Nf,numfmt="#,##0.00")
sc(j,"A13","ค่างานด่วน (เพิ่ม)"); sc(j,"B13",'=IF(AND(B11="volume",B8="ใช่",B10>=1,B10<50),IF(--B4>=1440,50,20)*B10,0)',Nf,numfmt="#,##0.00")
sc(j,"A15","ราคารวม (บาท)",B); sc(j,"B15",'=IF(ISNUMBER(B12),B12*B10+B13,"สอบถามราคากับฝ่ายขาย 081-770-7544")',Big,RES,numfmt="฿#,##0.00")
sc(j,"A16","ราคาต่อป้าย"); sc(j,"B16",'=IF(ISNUMBER(B12),(B12*B10+B13)/B7,"—")',Nf,numfmt="฿#,##0.00")
sc(j,"A18","หมายเหตุ: วัสดุที่เป็น “สอบถามราคา” ยังไม่มีเรตในระบบ — โทร 081-770-7544 / LINE @thaiprint.co.th",Sm); j.merge_cells("A18:D18")
for col,w in {"A":22,"B":30}.items(): j.column_dimensions[col].width=w

# ================= QUOTE (พร้อมพิมพ์ A4) =================
q=wb.create_sheet("Quote"); q.sheet_view.showGridLines=False
for col,w in {"A":22,"B":16,"C":12,"D":16,"E":16}.items(): q.column_dimensions[col].width=w
# logo + header
sc(q,"A1","TP",Logo,HFILL,align="center"); q.merge_cells("A1:A3"); q["A1"].alignment=Alignment(horizontal="center",vertical="center")
sc(q,"B1","บริษัท ไทยพริ้นท์ อินเตอร์กรุ๊ป จำกัด",CoName); q.merge_cells("B1:E1")
sc(q,"B2","โรงพิมพ์ฉลากสินค้าและกล่องบรรจุภัณฑ์ · โทร 02-149-4518, 081-770-7544",Sm); q.merge_cells("B2:E2")
sc(q,"B3","89/38 หมู่ 8 ต.เสาธงหิน อ.บางใหญ่ นนทบุรี 11140 · เลขผู้เสียภาษี 0125564010433",Sm); q.merge_cells("B3:E3")
sc(q,"A5","ใบเสนอราคา / QUOTATION",Font(name=TH,size=18,bold=True),SEC,align="center"); q.merge_cells("A5:E5")
sc(q,"A7","เลขที่เอกสาร"); sc(q,"B7",'="QT"&TEXT(TODAY(),"yymmdd")&"-001"',Nf,INP,border=BORD)
sc(q,"D7","วันที่"); sc(q,"E7","=TODAY()",Nf,border=BORD,numfmt="dd/mm/yyyy")
sc(q,"A8","ลูกค้า"); sc(q,"B8","",Nf,INP,border=BORD); q.merge_cells("B8:C8")
sc(q,"D8","โทร"); sc(q,"E8","",Nf,INP,border=BORD)
sc(q,"A9","ชื่องาน"); sc(q,"B9",'="งานพิมพ์ "&Offset!B4',Nf,border=BORD); q.merge_cells("B9:E9")
# spec line
sc(q,"A11",'=CONCATENATE("วัสดุ: ",Offset!B4," | ขนาด: ",Offset!B6," × ",Offset!B7," ซม. | สี หน้า/หลัง: ",Offset!B8,"/",Offset!B9," | เครื่อง: ",Offset!B41)',Sm); q.merge_cells("A11:E11")
# items table
for i,h in enumerate(["รายการ","จำนวน","หน่วย","ราคา/หน่วย","รวม (บาท)"]): sc(q,f"{'ABCDE'[i]}13",h,B,SEC,border=BORD,align="center")
sc(q,"A14","ค่างานพิมพ์ (เพลท+พิมพ์+วัสดุ)",Nf,border=BORD)
sc(q,"B14","=Offset!B11",Nf,border=BORD,align="center"); sc(q,"C14","ชิ้น",Nf,border=BORD,align="center")
sc(q,"D14","=Offset!B42/Offset!B11",Nf,border=BORD,numfmt="#,##0.00"); sc(q,"E14","=Offset!B42",Nf,border=BORD,numfmt="#,##0.00")
sc(q,"A15","ค่างานหลังพิมพ์",Nf,border=BORD); sc(q,"B15","",Nf,border=BORD); sc(q,"C15","",Nf,border=BORD); sc(q,"D15","",Nf,border=BORD)
sc(q,"E15","=Offset!B43",Nf,border=BORD,numfmt="#,##0.00")
for r in (16,17):
    for cLet in "ABCDE": sc(q,f"{cLet}{r}","",Nf,border=BORD)
sc(q,"D18","รวมเป็นเงิน",B,border=BORD,align="right"); sc(q,"E18","=Offset!B42+Offset!B43",Nf,border=BORD,numfmt="#,##0.00")
sc(q,"D19","VAT (Prices!B7 %)",Nf,border=BORD,align="right"); sc(q,"E19","=(Offset!B42+Offset!B43)*Prices!$B$7/100",Nf,border=BORD,numfmt="#,##0.00")
sc(q,"D20","ราคารวมทั้งสิ้น",B,RES,border=BORD,align="right"); sc(q,"E20","=(Offset!B42+Offset!B43)*(1+Prices!$B$7/100)",Big,RES,border=BORD,numfmt="฿#,##0.00")
sc(q,"A22","หมายเหตุ: ราคาประเมินเบื้องต้น ยืนราคา 15 วัน · ราคาอาจเปลี่ยนตามแบบจริง/งานตกแต่งพิเศษ",Sm); q.merge_cells("A22:E22")
sc(q,"B25","ผู้เสนอราคา",Nf,align="center"); q.merge_cells("B25:C25")
sc(q,"D25","ผู้อนุมัติ/ลูกค้า",Nf,align="center"); q.merge_cells("D25:E25")
sc(q,"B26","__________________",Nf,align="center"); q.merge_cells("B26:C26")
sc(q,"D26","__________________",Nf,align="center"); q.merge_cells("D26:E26")
# page setup A4 portrait
q.page_setup.orientation="portrait"; q.page_setup.paperSize=9  # A4
q.page_setup.fitToWidth=1; q.page_setup.fitToHeight=1; q.sheet_properties.pageSetUpPr=None
from openpyxl.worksheet.properties import PageSetupProperties
q.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
q.print_area="A1:E27"

wb.save("ThaiPrint-Calculator-Offline.xlsx")
print("saved")
