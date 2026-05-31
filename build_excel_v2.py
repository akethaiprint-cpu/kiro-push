"""
สร้าง Excel ระบบคำนวณราคาสิ่งพิมพ์ ThaiPrint v2
อัปเดตจาก v1 — รองรับ:
- 13 เครื่อง min/max ตาม datasheet
- Spoilage profiles ใหม่ 6 แบบ + Perfecting modifier
- Dual plate cost tables (legacy/datasheet2025)
- Paper Cutting Optimizer reference table
- Min/Max sheet validation (เตือนเมื่อกระดาษเล็ก/ใหญ่เกิน)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

wb = Workbook()

# ─── Styles ────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RESULT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
INFO_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

WHITE_BOLD = Font(name="TH SarabunPSK", size=14, bold=True, color="FFFFFF")
BOLD = Font(name="TH SarabunPSK", size=14, bold=True)
NORMAL = Font(name="TH SarabunPSK", size=14)
SMALL = Font(name="TH SarabunPSK", size=12)
LARGE_GREEN = Font(name="TH SarabunPSK", size=18, bold=True, color="006100")
LARGE_RED = Font(name="TH SarabunPSK", size=14, bold=True, color="9C0006")

THIN = Side(border_style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header(c):
    c.font = WHITE_BOLD; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

def style_section(c):
    c.font = BOLD; c.fill = SECTION_FILL; c.alignment = CENTER; c.border = BOX

def style_normal(c, bold=False, align="left"):
    c.font = BOLD if bold else NORMAL
    c.alignment = LEFT if align == "left" else (CENTER if align == "center" else RIGHT)
    c.border = BOX

def style_input(c):
    c.font = NORMAL; c.fill = INPUT_FILL; c.alignment = LEFT; c.border = BOX

def style_result(c):
    c.font = BOLD; c.fill = RESULT_FILL; c.alignment = RIGHT; c.border = BOX


# ============================================================
# SHEET 1: Machines (13 เครื่อง พร้อม min/max + plate cost dual)
# ============================================================
ws_m = wb.active
ws_m.title = "Machines"

ws_m["A1"] = "ตารางเครื่องพิมพ์ Heidelberg + Komori (13 รุ่น) — datasheet 2025"
ws_m.merge_cells("A1:I1")
style_header(ws_m["A1"])
ws_m.row_dimensions[1].height = 28

headers = ["Key", "ชื่อเครื่อง", "Min W (ซม.)", "Min H (ซม.)", "Max W (ซม.)", "Max H (ซม.)",
           "Plate Size", "ค่าเพลท Legacy", "ค่าเพลท Datasheet2025"]
for col, h in enumerate(headers, 1):
    c = ws_m.cell(row=2, column=col, value=h)
    style_section(c)

machines = [
    ("heidelberg_gto46", "Heidelberg GTO 46", 14.8, 10.5, 46, 34, "ตัด 8 (เล็ก)", 200, 150),
    ("heidelberg_sm52", "Heidelberg SM52", 14.5, 10.5, 52, 37, "ตัด 8 (กลาง)", 250, 150),
    ("heidelberg_gto52", "Heidelberg GTO 52", 18, 10.5, 52, 36, "ตัด 4", 250, 150),
    ("heidelberg_mo", "Heidelberg MO (MOZ/MOV)", 28, 21, 65, 48, "ตัด 4 (B2)", 300, 300),
    ("heidelberg_movp", "Heidelberg MOVP (Perfector)", 28, 21, 65, 48, "ตัด 4 (B2)", 300, 300),
    ("heidelberg_sx52", "Heidelberg SX52", 14.5, 10.5, 52, 37, "ตัด 2", 450, 150),
    ("heidelberg_sm74", "Heidelberg SM74", 28, 21, 74, 53, "ตัด 2 (B2+)", 500, 500),
    ("heidelberg_sx74", "Heidelberg SX74", 28, 21, 74, 53, "ตัด 2", 500, 500),
    ("heidelberg_xl75", "Heidelberg XL75", 35, 28, 74, 58.5, "ตัด 2", 550, 600),
    ("heidelberg_cd102", "Heidelberg CD102 / XL102", 48, 34, 102, 72, "ตัด 1 (B1)", 600, 600),
    ("heidelberg_sx102", "Heidelberg SX102", 48, 34, 102, 72, "ตัด 1", 650, 600),
    ("heidelberg_xl106", "Heidelberg XL106", 48, 34, 105, 75, "ตัด 1", 700, 1000),
    ("komori_s40", "Komori Lithrone S40", 48, 34, 106, 75, "ตัด 1 (B1)", 700, 1000),
]
for i, m in enumerate(machines, 3):
    for col, val in enumerate(m, 1):
        c = ws_m.cell(row=i, column=col, value=val)
        style_normal(c, align="center" if col in (3, 4, 5, 6, 8, 9) else "left")
        if col in (8, 9):
            c.number_format = "#,##0"

widths_m = [22, 28, 12, 12, 12, 12, 16, 16, 18]
for col, w in enumerate(widths_m, 1):
    ws_m.column_dimensions[get_column_letter(col)].width = w

# Note
ws_m["A17"] = ("หมายเหตุ: ตัด N = ขนาดเพลท/เครื่อง ไม่ใช่การตัดกระดาษ • "
               "MO ห้ามใช้กระดาษ 31×43 (ใหญ่เกิน) • ใช้ Paper Cutting เพื่อตัดกระดาษโรงงานให้พอดีเครื่องเล็กกว่า")
ws_m.merge_cells("A17:I17")
style_normal(ws_m["A17"], align="center")
ws_m["A17"].fill = INFO_FILL
ws_m["A17"].font = SMALL

# ============================================================
# SHEET 2: PrintCost (ตาราง Heidelberg)
# ============================================================
ws_pc = wb.create_sheet("PrintCost")

ws_pc["A1"] = "ตารางค่าพิมพ์ Heidelberg (ตัด 4 / ตัด 2 × Conv / UV)"
ws_pc.merge_cells("A1:E1")
style_header(ws_pc["A1"])
ws_pc.row_dimensions[1].height = 28

pc_headers = ["Plate Class", "Ink Type", "ช่วงใบพิมพ์", "เหมา (บาท/สี)", "ส่วนเกิน (บาท/ใบ/สี)"]
for col, h in enumerate(pc_headers, 1):
    style_section(ws_pc.cell(row=2, column=col, value=h))

pc_rows = [
    ("ตัด 4", "Conventional", "≤ 10,000 ใบ", 900, 0),
    ("ตัด 4", "Conventional", "> 10,000 ใบ", 900, 0.10),
    ("ตัด 4", "UV", "≤ 1,000 ใบ", 1500, 0),
    ("ตัด 4", "UV", "> 1,000 ใบ", 1500, 1.00),
    ("ตัด 2", "Conventional", "≤ 10,000 ใบ", 1200, 0),
    ("ตัด 2", "Conventional", "> 10,000 ใบ", 1200, 0.20),
    ("ตัด 2", "UV", "≤ 10,000 ใบ", 2000, 0),
    ("ตัด 2", "UV", "> 10,000 ใบ", 2000, 1.50),
]
for i, r in enumerate(pc_rows, 3):
    for col, val in enumerate(r, 1):
        c = ws_pc.cell(row=i, column=col, value=val)
        style_normal(c, align="center" if col in (1, 2, 3) else "right")
        if col in (4, 5) and isinstance(val, (int, float)):
            c.number_format = "#,##0.00"

ws_pc["A12"] = ("Sheetwise = ค่าพิมพ์ × 2 | Work-and-Turn = × 1 | "
                "ตัด 1 → ใช้อัตรา ตัด 2 (ประมาณ) | ตัด 8 → ใช้อัตรา ตัด 4 (ประมาณ)")
ws_pc.merge_cells("A12:E12")
style_normal(ws_pc["A12"], align="center")
ws_pc["A12"].fill = INFO_FILL

widths_pc = [16, 16, 22, 18, 22]
for col, w in enumerate(widths_pc, 1):
    ws_pc.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SHEET 3: Spoilage (Profiles 6 แบบ + Perfecting)
# ============================================================
ws_sp = wb.create_sheet("Spoilage")

ws_sp["A1"] = "Spoilage Profiles (6 แบบ) + Perfecting Modifier"
ws_sp.merge_cells("A1:D1")
style_header(ws_sp["A1"])
ws_sp.row_dimensions[1].height = 28

sp_headers = ["Job Type Key", "ชื่อ (ไทย)", "Rate (%)", "ขั้นต่ำ (ใบ)"]
for col, h in enumerate(sp_headers, 1):
    style_section(ws_sp.cell(row=2, column=col, value=h))

sp_rows = [
    ("oneColorRepeat", "1 สี ซ้ำ", 3, 50),
    ("twoColorGeneral", "2 สี ทั่วไป", 5, 100),
    ("fourColorRepeat", "4 สี ซ้ำ", 8, 100),
    ("fourColorFirst", "4 สี ครั้งแรก", 10, 150),
    ("uvOrPantone", "UV / Pantone", 10, 100),
    ("specialPaper", "กระดาษพิเศษ (สติ๊กเกอร์/PVC)", 15, 150),
]
for i, r in enumerate(sp_rows, 3):
    for col, val in enumerate(r, 1):
        c = ws_sp.cell(row=i, column=col, value=val)
        style_normal(c, align="center" if col in (3, 4) else "left")

# Perfecting section
ws_sp["A10"] = "Perfecting Modifier (เมื่อ printMethod = perfector)"
ws_sp.merge_cells("A10:D10")
style_section(ws_sp["A10"])

ws_sp["A11"] = "ค่า"; style_normal(ws_sp["A11"], bold=True)
ws_sp["B11"] = "Default"; style_normal(ws_sp["B11"], bold=True, align="center")
ws_sp["C11"] = "Min"; style_normal(ws_sp["C11"], bold=True, align="center")
ws_sp["D11"] = "Max"; style_normal(ws_sp["D11"], bold=True, align="center")

ws_sp["A12"] = "Perfecting +%"; style_normal(ws_sp["A12"])
ws_sp["B12"] = 2.5; style_normal(ws_sp["B12"], align="center")
ws_sp["C12"] = 2.0; style_normal(ws_sp["C12"], align="center")
ws_sp["D12"] = 3.0; style_normal(ws_sp["D12"], align="center")

ws_sp["A14"] = "สูตร: spoilageSheets = MAX(CEIL(minSheets × rate), profile.minSheets) | totalSheets = minSheets + spoilageSheets"
ws_sp.merge_cells("A14:D14")
ws_sp["A14"].fill = INFO_FILL
style_normal(ws_sp["A14"], align="center")
ws_sp["A14"].font = SMALL

widths_sp = [22, 32, 14, 16]
for col, w in enumerate(widths_sp, 1):
    ws_sp.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SHEET 4: PaperPrice (ราคากระดาษต่อ กก.)
# ============================================================
ws_pp = wb.create_sheet("PaperPrice")

ws_pp["A1"] = "ราคากระดาษต่อ กิโลกรัม (Charoen AkSorn พ.ค. 2569)"
ws_pp.merge_cells("A1:D1")
style_header(ws_pp["A1"])
ws_pp.row_dimensions[1].height = 28

pp_headers = ["ประเภทกระดาษ", "แกรม", "ยี่ห้อ/Brand", "ราคา/กก. (บาท)"]
for col, h in enumerate(pp_headers, 1):
    style_section(ws_pp.cell(row=2, column=col, value=h))

pp_rows = [
    ("ปอนด์ (Woodfree)", "60", "Sky Blue / Paper Plus", 31.00),
    ("ปอนด์ (Woodfree)", "70-120", "Sky Blue / Paper Plus", 30.00),
    ("ปอนด์ (Woodfree)", "60", "UPM PEFC", 32.00),
    ("ปอนด์ (Woodfree)", "70-80", "UPM PEFC", 31.00),
    ("ปอนด์ (Woodfree)", "60", "Premium X", 30.00),
    ("ปอนด์ (Woodfree)", "70-100", "Premium X", 29.00),
    ("อาร์ตมัน/ด้าน", "80-90", "Superkote", 34.00),
    ("อาร์ตมัน/ด้าน", "100-157", "Superkote", 32.00),
    ("อาร์ตมัน/ด้าน", "80", "Nevia Spakling", 33.00),
    ("อาร์ตมัน/ด้าน", "85-90", "Nevia Spakling", 32.00),
    ("อาร์ตมัน/ด้าน", "100-157", "Nevia Spakling", 31.00),
    ("อาร์ตมัน/ด้าน", "80-90", "YUKI/HI-KOTE/ART-TECH", 32.00),
    ("อาร์ตมัน/ด้าน", "100-128", "YUKI/HI-KOTE/ART-TECH", 31.00),
    ("อาร์ตการ์ด 2 หน้า", "190-260", "Superkote/HI-KOTE", 31.00),
    ("อาร์ตการ์ด 2 หน้า", "300-420", "Superkote/HI-KOTE", 30.00),
    ("อาร์ตการ์ด 1 หน้า (Ivory)", "190-300", "ฟอลเรท/UPM/PrimaArt", 32.00),
    ("อาร์ตการ์ด 1 หน้า (Ivory)", "330-380", "ฟอลเรท/UPM/PrimaArt", 31.00),
    ("กล่องแป้งหลังเทา", "250-450", "Royal/Lerthai", 22.00),
    ("คราฟท์", "50-170", "หลายยี่ห้อ", 28.00),
    ("กระดาษการ์ดขาว", "150-250", "หลายยี่ห้อ", 35.00),
]
for i, r in enumerate(pp_rows, 3):
    for col, val in enumerate(r, 1):
        c = ws_pp.cell(row=i, column=col, value=val)
        style_normal(c, align="left" if col in (1, 3) else ("center" if col == 2 else "right"))
        if col == 4:
            c.number_format = "#,##0.00"

widths_pp = [28, 14, 28, 18]
for col, w in enumerate(widths_pp, 1):
    ws_pp.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SHEET 5: PaperCutting (ตารางการตัดกระดาษ)
# ============================================================
ws_pcut = wb.create_sheet("PaperCutting")

ws_pcut["A1"] = "Paper Cutting Reference — กระดาษโรงงาน → ตัดเป็นชิ้นสำหรับเครื่องพิมพ์"
ws_pcut.merge_cells("A1:F1")
style_header(ws_pcut["A1"])
ws_pcut.row_dimensions[1].height = 28

pcut_headers = ["Factory Sheet", "ตัด (rows×cols)", "ชิ้น/ใบ", "ขนาดชิ้น (ซม.)", "เครื่องที่ใช้ได้", "งานที่เหมาะ"]
for col, h in enumerate(pcut_headers, 1):
    style_section(ws_pcut.cell(row=2, column=col, value=h))

pcut_rows = [
    ("31×43 นิ้ว", "(1,2)", 2, "78.74 × 54.61", "CD102, SX102, XL106, Komori S40", "A1+ โปสเตอร์ใหญ่"),
    ("31×43 นิ้ว", "(2,2)", 4, "54.61 × 39.37", "MO, MOVP, SM74, SX74, XL75, CD102, SX102, XL106, Komori", "A2+ ปก โปสเตอร์"),
    ("31×43 นิ้ว", "(4,2)", 8, "54.61 × 19.69", "ไม่มี (ชิ้นยาวเกินด้านสั้นเครื่อง)", "ต้องตัดแบบอื่น ⚠️"),
    ("31×43 นิ้ว", "(4,4)", 16, "27.31 × 19.69", "GTO46, GTO52, SM52, SX52", "A4 (เสียเศษมาก ⚠️)"),
    ("25×36 นิ้ว", "(1,1)", 1, "91.44 × 63.50", "CD102, SX102, XL106, Komori S40", "B1"),
    ("25×36 นิ้ว", "(2,2)", 4, "45.72 × 31.75", "GTO52, MO, MOVP, SM52, SX52, SM74, SX74, XL75", "A3+ โปสเตอร์ A3"),
    ("25×36 นิ้ว", "(3,3)", 9, "30.48 × 21.17", "GTO46, GTO52, SM52, SX52, MO, MOVP, SM74, SX74", "✅ A4 เสียเศษน้อยสุด"),
    ("25×36 นิ้ว", "(6,6)", 36, "15.24 × 10.58", "GTO46, SM52, SX52", "A6 นามบัตร"),
    ("24×35 นิ้ว", "(1,1)", 1, "88.90 × 60.96", "CD102, SX102, XL106, Komori S40", "B1-"),
    ("24×35 นิ้ว", "(2,2)", 4, "44.45 × 30.48", "GTO46, GTO52, SM52, SX52, MO, MOVP, SM74, SX74, XL75", "A3+"),
    ("24×35 นิ้ว", "(2,4)", 8, "44.45 × 15.24", "GTO46, GTO52, SM52, SX52", "งานยาว"),
    ("24×35 นิ้ว", "(8,4)", 32, "22.23 × 7.62", "ไม่มี (ชิ้นเล็ก/แคบเกินเครื่อง)", "งานเล็กพิเศษ ⚠️"),
    ("19×25 นิ้ว", "(1,1)", 1, "63.50 × 48.26", "SM74, SX74, XL75, CD102, SX102, XL106, Komori", "B2"),
    ("19×25 นิ้ว", "(2,2)", 4, "31.75 × 24.13", "GTO46, GTO52, SM52, SX52, MO, MOVP, SM74, SX74", "A4+"),
    ("19×25 นิ้ว", "(4,2)", 8, "31.75 × 12.07", "GTO46, GTO52, SM52, SX52", "งานยาว"),
]
for i, r in enumerate(pcut_rows, 3):
    for col, val in enumerate(r, 1):
        c = ws_pcut.cell(row=i, column=col, value=val)
        style_normal(c, align="center" if col in (1, 2, 3, 4) else "left")
        if col == 6 and isinstance(val, str) and "✅" in val:
            c.fill = RESULT_FILL
        elif col == 6 and isinstance(val, str) and "⚠️" in val:
            c.fill = WARN_FILL

widths_pcut = [16, 14, 10, 18, 36, 28]
for col, w in enumerate(widths_pcut, 1):
    ws_pcut.column_dimensions[get_column_letter(col)].width = w

# Note
ws_pcut["A19"] = "💡 งาน A4 → ใช้ 25×36 นิ้ว (3,3) ดีที่สุด — เสียเศษแค่ 5 มม. × 23.4 มม. | 24×35 (2,4) = 21×29.7 ก็เหมาะ ห้ามใช้ 31×43"
ws_pcut.merge_cells("A19:F19")
style_normal(ws_pcut["A19"], align="center")
ws_pcut["A19"].fill = INFO_FILL
ws_pcut["A19"].font = SMALL

# ============================================================
# SHEET 6: QuickSelect
# ============================================================
ws_qs = wb.create_sheet("QuickSelect")

ws_qs["A1"] = "Quick Select — แนะนำเครื่อง / วิธีพิมพ์ / จำนวนสี"
ws_qs.merge_cells("A1:C1")
style_header(ws_qs["A1"])
ws_qs.row_dimensions[1].height = 28

# Section: by Quantity
ws_qs["A3"] = "1. แนะนำวิธีพิมพ์ตามจำนวนสั่ง"
ws_qs.merge_cells("A3:C3")
style_section(ws_qs["A3"])

qs_q_headers = ["จำนวนสั่ง", "วิธีแนะนำ", "เหตุผล"]
for col, h in enumerate(qs_q_headers, 1):
    style_section(ws_qs.cell(row=4, column=col, value=h))

qs_qty = [
    ("< 500 ชิ้น", "Work-and-Turn", "ค่าเพลทประหยัดครึ่งหนึ่ง"),
    ("500–2,000 ชิ้น", "Work-and-Turn", "ยังประหยัดได้ชัดเจน"),
    ("2,001–5,000 ชิ้น", "Work-and-Turn / Sheetwise", "พิจารณาตามคุณภาพ"),
    ("> 5,000 ชิ้น", "Sheetwise", "Registration ดีกว่า"),
]
for i, r in enumerate(qs_qty, 5):
    for col, val in enumerate(r, 1):
        style_normal(ws_qs.cell(row=i, column=col, value=val))

# Section: by Color Count
ws_qs["A11"] = "2. แนะนำเครื่องตามจำนวนสี"
ws_qs.merge_cells("A11:C11")
style_section(ws_qs["A11"])

qs_color_headers = ["งาน", "เครื่องที่แนะนำ", "หมายเหตุ"]
for col, h in enumerate(qs_color_headers, 1):
    style_section(ws_qs.cell(row=12, column=col, value=h))

qs_color = [
    ("1–2 สี", "GTO52, MO", "—"),
    ("4 สี (CMYK)", "GTO52, MO, SM74, XL106", "—"),
    ("5+ สี + Pantone/Coating", "GTO52, MO, SM74", "5+ unit"),
    ("พิมพ์ 2 ด้านรอบเดียว", "MOVP, SM74-2P, XL106-2P", "Perfector"),
]
for i, r in enumerate(qs_color, 13):
    for col, val in enumerate(r, 1):
        style_normal(ws_qs.cell(row=i, column=col, value=val))

# Section: by Paper Size
ws_qs["A19"] = "3. แนะนำเครื่องตามขนาดกระดาษ"
ws_qs.merge_cells("A19:C19")
style_section(ws_qs["A19"])

qs_size_headers = ["กระดาษ", "เครื่องที่รองรับ", "หมายเหตุ"]
for col, h in enumerate(qs_size_headers, 1):
    style_section(ws_qs.cell(row=20, column=col, value=h))

qs_size = [
    ("≤ 36×52 ซม. (A3+)", "GTO 46, GTO 52, SM52", "งานเล็ก ปริมาณน้อย"),
    ("≤ 48×65 ซม.", "MO / MOVP", "B2 นิยมในไทย (ตัด 4)"),
    ("≤ 52×74 ซม.", "SX52", "—"),
    ("24\"×35\" หรือ 25\"×36\"", "SM74 / CD102 / XL106", "เหมาะงาน A4 ไม่เสียเศษ"),
    ("≤ 53×74 ซม.", "SM74 / SX74", "B2+ สมรรถนะสูง"),
    ("≤ 53×75 ซม.", "XL75", "B2+ XL series"),
    ("≤ 72×102 ซม.", "CD102 / SX102", "B1"),
    ("≤ 75×106 ซม.", "XL106 / Komori S40", "B1 สมรรถนะสูงสุด"),
]
for i, r in enumerate(qs_size, 21):
    for col, val in enumerate(r, 1):
        style_normal(ws_qs.cell(row=i, column=col, value=val))

# A4 Warning
ws_qs["A30"] = "⚠️ งาน A4 ห้ามใช้กระดาษ 31×43 (เสียเศษมาก) — ใช้ 24×35 หรือ 25×36 แทน"
ws_qs.merge_cells("A30:C30")
style_normal(ws_qs["A30"], align="center")
ws_qs["A30"].fill = WARN_FILL
ws_qs["A30"].font = BOLD

widths_qs = [28, 36, 28]
for col, w in enumerate(widths_qs, 1):
    ws_qs.column_dimensions[get_column_letter(col)].width = w

print("Sheets 1-6 done. Saving partial workbook...")
wb.save("/Users/thaiprint/kiro/ThaiPrint-Pricing-Calculator-v2.xlsx")
print("Saved.")


# ============================================================
# SHEET 7: Calculator (กรอกข้อมูล + คำนวณราคา)
# ============================================================
ws_c = wb.create_sheet("Calculator", 0)  # insert at beginning

ws_c["A1"] = "ระบบคำนวณราคาใบพิมพ์ — ThaiPrint v2 (Datasheet 2025)"
ws_c.merge_cells("A1:E1")
style_header(ws_c["A1"])
ws_c.row_dimensions[1].height = 32

ws_c["A2"] = "🟡 ช่องเหลือง = กรอก  |  🟢 ช่องเขียว = ผลคำนวณ  |  🔴 = เตือน  |  Datasheet 2025: 13 เครื่อง + Spoilage 6 แบบ"
ws_c.merge_cells("A2:E2")
ws_c["A2"].font = SMALL
ws_c["A2"].alignment = CENTER
ws_c["A2"].fill = INFO_FILL

# ── Section: Input ──────────────────────────────────────────
ws_c["A4"] = "ข้อมูลงาน (กรอกที่ช่องสีเหลือง)"
ws_c.merge_cells("A4:E4")
style_section(ws_c["A4"])

input_rows = [
    (5, "ชื่องาน", "โบรชัวร์ A4 4 สี"),
    (6, "ขนาดชิ้นงาน — กว้าง (ซม.)", 21),
    (7, "ขนาดชิ้นงาน — ยาว (ซม.)", 29.7),
    (8, "ขนาดกระดาษ (ใส่เครื่องโดยตรง)", "24x35"),  # dropdown
    (9, "ชนิดกระดาษ", "อาร์ตมัน/ด้าน"),
    (10, "แกรม", 128),
    (11, "ราคากระดาษ (บาท/กก.)", 32.00),
    (12, "เครื่องพิมพ์ (Key)", "heidelberg_cd102"),  # dropdown
    (13, "จำนวนพิมพ์ (ชิ้น)", 1000),
    (14, "จำนวนสี", 4),
    (15, "ประเภทงาน (Spoilage)", "fourColorRepeat"),  # dropdown
    (16, "ประเภทหมึก", "Conventional"),  # dropdown
    (17, "วิธีพิมพ์", "Sheetwise"),  # dropdown
    (18, "Plate Cost Mode", "Legacy"),  # dropdown legacy/datasheet2025
    (19, "Gripper Margin (มม., 10–15)", 12),
    (20, "Side Lay (มม., 5–10)", 7),
    (21, "Perfecting Modifier (%, 2–3, ใช้เมื่อ Perfector)", 2.5),
]
for row, label, default in input_rows:
    lbl = ws_c.cell(row=row, column=1, value=label)
    style_normal(lbl, bold=True)
    inp = ws_c.cell(row=row, column=2, value=default)
    style_input(inp)
    ws_c.merge_cells(f"B{row}:E{row}")

# Dropdowns
dv_sheet = DataValidation(type="list", formula1='"31x43,25x36,24x35,19x25,custom"', allow_blank=False)
dv_sheet.add("B8"); ws_c.add_data_validation(dv_sheet)

machine_keys = ",".join([m[0] for m in machines])
dv_machine = DataValidation(type="list", formula1=f'"{machine_keys}"', allow_blank=False)
dv_machine.add("B12"); ws_c.add_data_validation(dv_machine)

dv_job = DataValidation(type="list",
    formula1='"oneColorRepeat,twoColorGeneral,fourColorRepeat,fourColorFirst,uvOrPantone,specialPaper"', allow_blank=False)
dv_job.add("B15"); ws_c.add_data_validation(dv_job)

dv_ink = DataValidation(type="list", formula1='"Conventional,UV"', allow_blank=False)
dv_ink.add("B16"); ws_c.add_data_validation(dv_ink)

dv_method = DataValidation(type="list",
    formula1='"Single,Sheetwise,Work-and-Turn,Work-and-Tumble,Perfector"', allow_blank=False)
dv_method.add("B17"); ws_c.add_data_validation(dv_method)

dv_plate = DataValidation(type="list", formula1='"Legacy,Datasheet2025"', allow_blank=False)
dv_plate.add("B18"); ws_c.add_data_validation(dv_plate)

# ── Section: Lookup ─────────────────────────────────────────
ws_c["A23"] = "ค่าที่ค้นจากตาราง (อัตโนมัติ)"
ws_c.merge_cells("A23:E23")
style_section(ws_c["A23"])

lookup_rows = [
    (24, "Sheet Width (ซม.)",
     '=IF(B8="31x43",78.74,IF(B8="25x36",63.5,IF(B8="24x35",60.96,IF(B8="19x25",48.26,B6))))'),
    (25, "Sheet Height (ซม.)",
     '=IF(B8="31x43",109.22,IF(B8="25x36",91.44,IF(B8="24x35",88.9,IF(B8="19x25",63.5,B7))))'),
    (26, "Machine Name",
     '=VLOOKUP(B12,Machines!A3:I15,2,FALSE)'),
    (27, "Machine Min W (ซม.)",
     '=VLOOKUP(B12,Machines!A3:I15,3,FALSE)'),
    (28, "Machine Min H (ซม.)",
     '=VLOOKUP(B12,Machines!A3:I15,4,FALSE)'),
    (29, "Machine Max W (ซม.)",
     '=VLOOKUP(B12,Machines!A3:I15,5,FALSE)'),
    (30, "Machine Max H (ซม.)",
     '=VLOOKUP(B12,Machines!A3:I15,6,FALSE)'),
    (31, "Plate Size",
     '=VLOOKUP(B12,Machines!A3:I15,7,FALSE)'),
    (32, "ค่าเพลท/สี (เลือก mode)",
     '=IF(B18="Datasheet2025",VLOOKUP(B12,Machines!A3:I15,9,FALSE),VLOOKUP(B12,Machines!A3:I15,8,FALSE))'),
    (33, "Plate Class (normalized)",
     '=IF(ISNUMBER(SEARCH("ตัด 4",B31)),"ตัด 4",IF(ISNUMBER(SEARCH("ตัด 2",B31)),"ตัด 2",IF(ISNUMBER(SEARCH("ตัด 1",B31)),"ตัด 2",IF(ISNUMBER(SEARCH("ตัด 8",B31)),"ตัด 4","ตัด 4"))))'),
    (34, "Spoilage Rate (%)",
     '=IF(B15="oneColorRepeat",3,IF(B15="twoColorGeneral",5,IF(B15="fourColorRepeat",8,IF(B15="fourColorFirst",10,IF(B15="uvOrPantone",10,IF(B15="specialPaper",15,5))))))'),
    (35, "Spoilage Min (ใบ)",
     '=IF(B15="oneColorRepeat",50,IF(B15="twoColorGeneral",100,IF(B15="fourColorRepeat",100,IF(B15="fourColorFirst",150,IF(B15="uvOrPantone",100,IF(B15="specialPaper",150,100))))))'),
    (36, "Effective Rate (+Perfecting)",
     '=B34+IF(B17="Perfector",B21,0)'),
    (37, "ตรวจกระดาษเข้าเครื่อง",
     '=IF(OR(AND(B24>=B27,B24<=B29,B25>=B28,B25<=B30),AND(B25>=B27,B25<=B29,B24>=B28,B24<=B30)),"✅ เข้าได้",IF(OR(B24>B29,B25>B29),"⚠️ ใหญ่เกิน","⚠️ เล็กเกิน"))'),
    (38, "หมายเหตุ Print Cost",
     '=IF(OR(ISNUMBER(SEARCH("ตัด 1",B31)),ISNUMBER(SEARCH("ตัด 8",B31))),"⚠️ ใช้อัตรา "&B33&" (ประมาณ)","")'),
]
for row, label, formula in lookup_rows:
    lbl = ws_c.cell(row=row, column=1, value=label)
    style_normal(lbl)
    f = ws_c.cell(row=row, column=2, value=formula)
    style_normal(f, align="left")
    ws_c.merge_cells(f"B{row}:E{row}")

# ── Section: Calculation ─────────────────────────────────────
ws_c["A40"] = "ผลคำนวณ (อัตโนมัติ)"
ws_c.merge_cells("A40:E40")
style_section(ws_c["A40"])

calc_rows = [
    (41, "Gripper (ซม.)", '=B19/10', "0.00"),
    (42, "Side Lay (ซม.)", '=B20/10', "0.00"),
    (43, "Bleed รวม (ซม.)", '=0.6', "0.00"),
    (44, "พื้นที่พิมพ์ — กว้าง", '=B24-B41', "0.00"),
    (45, "พื้นที่พิมพ์ — ยาว", '=B25-B42', "0.00"),
    (46, "ขนาดชิ้น+bleed (กว้าง)", '=B6+B43', "0.00"),
    (47, "ขนาดชิ้น+bleed (ยาว)", '=B7+B43', "0.00"),
    (48, "ชิ้น/ใบ — ปกติ", '=INT(B44/B46)*INT(B45/B47)', "#,##0"),
    (49, "ชิ้น/ใบ — หมุน 90°", '=INT(B44/B47)*INT(B45/B46)', "#,##0"),
    (50, "ชิ้น/ใบพิมพ์ (เลือกค่ามาก)", '=MAX(B48,B49)', "#,##0"),
    (51, "ใบพิมพ์ขั้นต่ำ", '=CEILING(B13/B50,1)', "#,##0"),
    (52, "Spoilage แบบคำนวณ %", '=CEILING(B51*B36/100,1)', "#,##0"),
    (53, "Spoilage จริง (ขั้นต่ำ Profile)", '=MAX(B52,B35)', "#,##0"),
    (54, "หมายเหตุ Spoilage",
     '=IF(B52<B35,"⚠️ ใช้ขั้นต่ำ "&B35&" ใบ ตาม profile","")', ""),
    (55, "ใบพิมพ์รวม (Total Sheets)", '=B51+B53', "#,##0"),
    # Plate
    (56, "Plate Sets", '=IF(B17="Sheetwise",2,1)', "#,##0"),
    (57, "ค่าเพลทรวม", '=B32*B14*B56', "#,##0.00"),
    # Print cost
    (58, "Flat Rate (เหมา)",
     '=IF(B33="ตัด 4",IF(B16="UV",1500,900),IF(B16="UV",2000,1200))', "#,##0"),
    (59, "Tier ขอบบน",
     '=IF(B33="ตัด 4",IF(B16="UV",1000,10000),10000)', "#,##0"),
    (60, "Overage (บาท/ใบ/สี)",
     '=IF(B33="ตัด 4",IF(B16="UV",1,0.10),IF(B16="UV",1.5,0.20))', "0.00"),
    (61, "ส่วนเกิน (ใบ)", '=MAX(B55-B59,0)', "#,##0"),
    (62, "ค่าพิมพ์ก่อน multiplier", '=B58*B14+B61*B60*B14', "#,##0.00"),
    (63, "Print Method × (Sheetwise=2)", '=IF(B17="Sheetwise",2,1)', "#,##0"),
    (64, "ค่าพิมพ์รวม", '=B62*B63', "#,##0.00"),
    # Paper
    (65, "น้ำหนักต่อใบ (กรัม)", '=(B24*B25*B10)/10000', "#,##0.00"),
    (66, "น้ำหนักรวม (กก.)", '=B65*B55/1000', "#,##0.00"),
    (67, "ต้นทุนกระดาษ", '=B66*B11', "#,##0.00"),
    (68, "ราคาขายกระดาษ (+20%)", '=B67*1.2', "#,##0.00"),
]
for row, label, formula, fmt in calc_rows:
    lbl = ws_c.cell(row=row, column=1, value=label)
    style_normal(lbl)
    f = ws_c.cell(row=row, column=2, value=formula)
    style_normal(f, align="right")
    if fmt:
        f.number_format = fmt
    ws_c.merge_cells(f"B{row}:E{row}")

# Total
ws_c["A70"] = "สรุปราคารวม"
ws_c.merge_cells("A70:E70")
style_section(ws_c["A70"])

ws_c["A71"] = "ราคารวมทั้งหมด (ค่าเพลท + ค่ากระดาษขาย + ค่าพิมพ์)"
ws_c["A71"].font = BOLD; ws_c["A71"].alignment = LEFT; ws_c["A71"].border = BOX
total_cell = ws_c["B71"]
total_cell.value = '=B57+B68+B64'
total_cell.number_format = '"฿"#,##0.00'
style_result(total_cell)
total_cell.font = LARGE_GREEN
ws_c.merge_cells("B71:E71")
ws_c.row_dimensions[71].height = 32

ws_c["A72"] = "ราคาต่อชิ้น"
ws_c["A72"].font = BOLD; ws_c["A72"].alignment = LEFT; ws_c["A72"].border = BOX
unit_cell = ws_c["B72"]
unit_cell.value = '=B71/B13'
unit_cell.number_format = '"฿"#,##0.00'
style_result(unit_cell)
ws_c.merge_cells("B72:E72")

# Quick Select hints
ws_c["A74"] = "💡 คำแนะนำ Quick Select"
ws_c.merge_cells("A74:E74")
style_section(ws_c["A74"])

ws_c["A75"] = "วิธีพิมพ์แนะนำ"
style_normal(ws_c["A75"])
ws_c["B75"] = ('=IF(B13<500,"Work-and-Turn — จำนวนน้อย ประหยัดเพลท",'
               'IF(B13<=2000,"Work-and-Turn — ยังประหยัดได้ชัดเจน",'
               'IF(B13<=5000,"Work-and-Turn / Sheetwise — พิจารณาตามคุณภาพ",'
               '"Sheetwise — Registration ดีกว่า")))')
style_normal(ws_c["B75"], align="left")
ws_c.merge_cells("B75:E75")

ws_c["A76"] = "เตือน A4 บน 31×43"
style_normal(ws_c["A76"])
ws_c["B76"] = ('=IF(AND(B8="31x43",ABS(B6-21)<=1,ABS(B7-29.7)<=1),'
               '"⚠️ งาน A4 ไม่ควรใช้ 31×43 — แนะนำ 24×35 หรือ 25×36 (ตัด 8 หรือ 9 ชิ้น)","")')
style_normal(ws_c["B76"], align="left")
ws_c.merge_cells("B76:E76")

# Column widths
widths_c = [42, 22, 22, 22, 22]
for col, w in enumerate(widths_c, 1):
    ws_c.column_dimensions[get_column_letter(col)].width = w

# Conditional formatting
ws_c.conditional_formatting.add("B37",
    FormulaRule(formula=['ISNUMBER(SEARCH("⚠️",B37))'], fill=WARN_FILL))
ws_c.conditional_formatting.add("B38",
    FormulaRule(formula=['ISNUMBER(SEARCH("⚠️",B38))'], fill=WARN_FILL))
ws_c.conditional_formatting.add("B54",
    FormulaRule(formula=['ISNUMBER(SEARCH("⚠️",B54))'], fill=WARN_FILL))
ws_c.conditional_formatting.add("B76",
    FormulaRule(formula=['ISNUMBER(SEARCH("⚠️",B76))'], fill=WARN_FILL))

# ============================================================
# SHEET 8: Quotation
# ============================================================
ws_q = wb.create_sheet("Quotation")

# Header
ws_q["A1"] = "บริษัท ไทยพริ้นท์ อินเตอร์กรุ๊ป จำกัด"
ws_q.merge_cells("A1:E1")
ws_q["A1"].font = Font(name="TH SarabunPSK", size=20, bold=True, color="1F4E79")
ws_q["A1"].alignment = CENTER
ws_q.row_dimensions[1].height = 30

ws_q["A2"] = "ThaiPrint Intergroup Co., Ltd."
ws_q.merge_cells("A2:E2")
ws_q["A2"].font = Font(name="TH SarabunPSK", size=12, italic=True)
ws_q["A2"].alignment = CENTER

ws_q["A4"] = "ใบเสนอราคา / QUOTATION"
ws_q.merge_cells("A4:E4")
ws_q["A4"].font = Font(name="TH SarabunPSK", size=18, bold=True)
ws_q["A4"].alignment = CENTER
ws_q["A4"].fill = SECTION_FILL
ws_q.row_dimensions[4].height = 28

# Customer info
ws_q["A6"] = "ลูกค้า:"; ws_q["A6"].font = BOLD
ws_q["B6"] = "________________________"; ws_q.merge_cells("B6:C6")
ws_q["D6"] = "วันที่:"; ws_q["D6"].font = BOLD
ws_q["E6"] = '=TODAY()'; ws_q["E6"].number_format = "dd/mm/yyyy"

ws_q["A7"] = "ที่อยู่:"; ws_q["A7"].font = BOLD
ws_q["B7"] = "________________________"; ws_q.merge_cells("B7:C7")
ws_q["D7"] = "เลขที่:"; ws_q["D7"].font = BOLD
ws_q["E7"] = "QT-2026-_____"

ws_q["A8"] = "โทร:"; ws_q["A8"].font = BOLD
ws_q["B8"] = "________________________"; ws_q.merge_cells("B8:C8")

# Job spec table
ws_q["A10"] = "รายละเอียดงาน"
ws_q.merge_cells("A10:E10")
style_section(ws_q["A10"])

job_info = [
    ("ชื่องาน", "=Calculator!B5"),
    ("ขนาดชิ้นงาน", '="กว้าง "&Calculator!B6&" × ยาว "&Calculator!B7&" ซม."'),
    ("จำนวน", '=Calculator!B13&" ชิ้น"'),
    ("จำนวนสี", '=Calculator!B14&" สี"'),
    ("ประเภทกระดาษ", '=Calculator!B9&" "&Calculator!B10&" แกรม"'),
    ("ขนาดกระดาษ", '=Calculator!B8&" (ใส่เครื่อง)"'),
    ("เครื่องพิมพ์", '=Calculator!B26'),
    ("ประเภทหมึก", '=Calculator!B16'),
    ("วิธีพิมพ์", '=Calculator!B17'),
    ("ประเภทงาน", '=Calculator!B15'),
]
for i, (label, formula) in enumerate(job_info, 11):
    lbl = ws_q.cell(row=i, column=1, value=label)
    style_normal(lbl, bold=True)
    val = ws_q.cell(row=i, column=2, value=formula)
    style_normal(val, align="left")
    ws_q.merge_cells(f"B{i}:E{i}")

# Cost breakdown table
ws_q["A22"] = "รายการต้นทุน"
ws_q.merge_cells("A22:E22")
style_section(ws_q["A22"])

cost_headers = ["ลำดับ", "รายการ", "จำนวน", "หน่วย", "ราคา (บาท)"]
for col, h in enumerate(cost_headers, 1):
    style_section(ws_q.cell(row=23, column=col, value=h))

cost_items = [
    ("1", "ค่าเพลทพิมพ์", '=Calculator!B14*Calculator!B56', "แผ่น", '=Calculator!B57'),
    ("2", "ค่ากระดาษ (รวม markup 20%)", '=Calculator!B55', "ใบพิมพ์", '=Calculator!B68'),
    ("3", "ค่าพิมพ์", '=Calculator!B55', "ใบพิมพ์", '=Calculator!B64'),
]
for i, (no, item, qty, unit, price) in enumerate(cost_items, 24):
    style_normal(ws_q.cell(row=i, column=1, value=no), align="center")
    style_normal(ws_q.cell(row=i, column=2, value=item))
    qcell = ws_q.cell(row=i, column=3, value=qty); qcell.number_format = "#,##0"; style_normal(qcell, align="right")
    style_normal(ws_q.cell(row=i, column=4, value=unit), align="center")
    pcell = ws_q.cell(row=i, column=5, value=price); pcell.number_format = '#,##0.00'; style_normal(pcell, align="right")

# Subtotal/Total
ws_q.cell(row=28, column=1).border = BOX
sub_lbl = ws_q.cell(row=28, column=2, value="ยอดรวมต้นทุน")
sub_lbl.font = BOLD; sub_lbl.alignment = RIGHT; sub_lbl.border = BOX
ws_q.merge_cells("B28:D28")
sub_val = ws_q.cell(row=28, column=5, value='=SUM(E24:E26)')
sub_val.number_format = '#,##0.00'; sub_val.font = BOLD; sub_val.alignment = RIGHT; sub_val.border = BOX

tot_lbl = ws_q.cell(row=29, column=2, value="ราคาเสนอ (รวมทั้งหมด)")
tot_lbl.font = Font(name="TH SarabunPSK", size=16, bold=True); tot_lbl.alignment = RIGHT; tot_lbl.border = BOX
ws_q.merge_cells("B29:D29")
tot_val = ws_q.cell(row=29, column=5, value='=Calculator!B71')
tot_val.number_format = '"฿"#,##0.00'
tot_val.font = Font(name="TH SarabunPSK", size=16, bold=True, color="006100")
tot_val.fill = RESULT_FILL; tot_val.alignment = RIGHT; tot_val.border = BOX
ws_q.row_dimensions[29].height = 30

unit_lbl = ws_q.cell(row=30, column=2, value="ราคาต่อชิ้น")
unit_lbl.font = BOLD; unit_lbl.alignment = RIGHT; unit_lbl.border = BOX
ws_q.merge_cells("B30:D30")
unit_val = ws_q.cell(row=30, column=5, value='=Calculator!B72')
unit_val.number_format = '"฿"#,##0.00'; unit_val.font = BOLD; unit_val.alignment = RIGHT; unit_val.border = BOX

# Notes
ws_q["A32"] = "หมายเหตุ:"; ws_q["A32"].font = BOLD
ws_q["A33"] = "• ราคานี้ยังไม่รวม VAT 7%"
ws_q["A34"] = "• ราคาเสนอมีผลภายใน 30 วัน"
ws_q["A35"] = "• ระยะเวลาส่งมอบ 7-14 วันทำการหลังยืนยันแบบ"
for r in range(32, 36):
    ws_q.cell(row=r, column=1).font = NORMAL
    ws_q.merge_cells(f"A{r}:E{r}")

ws_q["A37"] = "ผู้เสนอราคา"; ws_q["A37"].font = BOLD
ws_q["A38"] = "_____________________"
ws_q["D37"] = "ผู้อนุมัติ"; ws_q["D37"].font = BOLD
ws_q["D38"] = "_____________________"

widths_q = [10, 30, 12, 12, 18]
for col, w in enumerate(widths_q, 1):
    ws_q.column_dimensions[get_column_letter(col)].width = w

ws_q.page_setup.orientation = ws_q.ORIENTATION_PORTRAIT
ws_q.page_setup.paperSize = ws_q.PAPERSIZE_A4
ws_q.page_setup.fitToWidth = 1
ws_q.page_setup.fitToHeight = 1
ws_q.sheet_properties.pageSetUpPr.fitToPage = True
ws_q.print_options.horizontalCentered = True
ws_q.page_margins.left = 0.4
ws_q.page_margins.right = 0.4
ws_q.page_margins.top = 0.5
ws_q.page_margins.bottom = 0.5

# Reorder sheets
desired = ["Calculator", "Quotation", "Machines", "PrintCost", "Spoilage",
           "PaperPrice", "PaperCutting", "QuickSelect"]
for i, name in enumerate(desired):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        if idx != i:
            wb.move_sheet(name, offset=i - idx)

wb.save("/Users/thaiprint/kiro/ThaiPrint-Pricing-Calculator-v2.xlsx")
print("✅ All 8 sheets created!")
print(f"Order: {wb.sheetnames}")
